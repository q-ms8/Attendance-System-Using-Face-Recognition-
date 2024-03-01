from datetime import datetime
import os
import cv2
import numpy as np
import face_recognition
import pandas as pd
import openpyxl


def read_Images(path):
    images = []  # empty list will contain all images in the images file as a matrix of pixels
    myList = os.listdir(path)  # list of all directory files in the 'path'

    for cl in myList:
        curImg = cv2.imread(f'{path}/{cl}')
        images.append(curImg)
    return images


def student_name(file_path):
    df = pd.read_excel(file_path)
    students = df.iloc[1:, 1].tolist()
    return students


def findEncodings(images):
    """
    :param images:
    :return: list of the encoding for a list of images
    """
    encodeList = []
    for img in images:
        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        encode = face_recognition.face_encodings(img)[0]
        encodeList.append(encode)
    return encodeList


def write_time_to_excel(file_path, column=3):

    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Get the current time
    current_time = datetime.now()

    # Write the current time to the specified cell
    sheet.cell(row=1, column=column, value=current_time)
    workbook.save(file_path)


def write_date_to_excel(file_path, column=3):

    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Get the current date
    current_date = datetime.now().date()

    # Write the current date to the specified cell
    sheet.cell(row=2, column=column, value=current_date)
    workbook.save(file_path)


def markAttendance(name, xlsx_file):
    # Load the workbook and sheet
    workbook = openpyxl.load_workbook(xlsx_file)
    sheet = workbook.active

    # Get the current date (assuming today's date)
    current_date = datetime.now().date()

    # Find the row corresponding to the student's name
    for row in range(2, sheet.max_row + 1):
        student_name_cell = sheet.cell(row=row, column=2)
        if student_name_cell.value == name:
            break

    # If the student's name is found, mark attendance
    if student_name_cell.value == name:
        # Check if attendance has already been marked for the current date
        attendance_cell = sheet.cell(row=row, column=current_date.weekday() + 1)
        if attendance_cell.value is None:
            # Mark attendance by writing 'P' in the corresponding cell
            attendance_cell.value = 'P'
            workbook.save(xlsx_file)
            print(f"Attendance marked for {name} on {current_date}")
        else:
            print(f"Attendance already marked for {name} on {current_date}")
    else:
        print(f"Student '{name}' not found in the attendance sheet")


xlsx_file = 'attendance_sheet.xlsx'
path = 'ImagesAttendance'  # the path of the file that contain the images
images = read_Images(path)
encodListKnown = findEncodings(images)
cap = cv2.VideoCapture(0)
className = student_name(xlsx_file)


while True:
    success, img = cap.read()
    imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
    imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)

    facesCurFrame = face_recognition.face_locations(imgS)
    encodesCurFrame = face_recognition.face_encodings(imgS, facesCurFrame)
    write_time_to_excel(xlsx_file)
    write_date_to_excel(xlsx_file)

    for encodeFace, faceLoc in zip(encodesCurFrame, facesCurFrame):
        matches = face_recognition.compare_faces(encodListKnown, encodeFace)
        faceDis = face_recognition.face_distance(encodListKnown, encodeFace)
        matchIndex = np.argmin(faceDis)

        if matches[matchIndex]:
            name = className[matchIndex]
            y1, x2, y2, x1 = faceLoc
            y1, x2, y2, x1 = 4*y1, 4*x2, 4*y2, 4*x1
            cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
            cv2.rectangle(img, (x1, y2-35), (x2, y2,), (0, 255, 0), cv2.FILLED)
            cv2.putText(img, name, (x1+6, y2-6), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)
            markAttendance(name, xlsx_file)

    cv2.imshow('Webcam', img)
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break